from fi import txt #导入文件库
from openpyxl import Workbook #excel 
import requests #爬虫库
import re #正则表达式 
#import random
import time #时间控制 

wb = Workbook() #创建工作簿
ws = wb.active #获取活跃的工作簿
print('根目录下关键词.txt内容修改即可，速度默认为0.5秒')
resultpath='关键词.txt' #定义待查的关键词
lineslist=txt.ReadTxtName(resultpath) #使用函数传入变量
#print(lineslist)



url = "http://www.baidu.com/s?rsv_bp=1&rsv_idx=1&tn=baidu&wd=intitle%3A" #定义url信息


headerss = {
    'Accept': 'application/json, text/javascript, */*; q=0.01',
    'Accept-Encoding': 'gzip, deflate, br',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Connection': 'Keep-Alive',
    'User-Agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Mobile Safari/537.36',
    'sec-ch-ua': '"Google Chrome";v="87", " Not;A Brand";v="99", "Chromium";v="87"'
}


hen = 1
for i in lineslist:
    data = ws['A%d' % hen] = i #写入表格 的A列
    hou = '&usm=3&rsv_idx=2&rsv_page=1' #玄学
    fullurl = url + i + hou
    s = requests.Session()
    html = s.get(fullurl,timeout=5,headers=headerss) 
    htmlutf = html.content   #转码步骤一
    html_doc=str(htmlutf,'utf-8') #html_doc=html.decode("utf-8","ignore")   #转码步骤二 转换成utf8
    urls = re.findall('<span class="nums_text">百度为您找到相关结果约(.*?)个</span>',html_doc)   #<span class="nums_text">百度为您找到相关结果约(.*?)个</span>
    print(urls)
    for x in urls:
        data = ws['B%d' % hen] = x
        print('已添加'+i+x)
    hen += 1   #加一个值
    #print(hen)
    #time.sleep(0.5)
wb.save('百度.xlsx')