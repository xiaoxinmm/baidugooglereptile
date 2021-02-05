from fi import txt
from openpyxl import Workbook
import requests
import re
import time

wb = Workbook() #创建工作簿
ws = wb.active #获取活跃的工作簿
print('根目录下关键词.txt内容修改即可，速度默认为1秒')

resultpath='关键词.txt' #定义待查的关键词
lineslist=txt.ReadTxtName(resultpath) #使用函数传入变量
#print(lineslist)


#一下三个url全是中文的
url = "http://www.google.com/search?hl=zh-CN&sxsrf=ALeKk01r6JJPiIY4JiZgQbtpdmirZ3pdrA%3A1612419427381&source=hp&ei=Y5EbYJi7FO6W4-EP2tKn-AI&q=intitle%3A" #定义url信息
url2 = '&oq=intitle%3A'
url3 =  '&gs_lcp=CgZwc3ktYWIQAzoHCCMQ6gIQJ1C2Nli2NmDxUGgCcAB4AIABiAKIAYgCkgEDMi0xmAEAoAECoAEBqgEHZ3dzLXdperABCg&sclient=psy-ab&ved=0ahUKEwjY-Me5ys_uAhVuyzgGHVrpCS8Q4dUDCAc&uact=5'


#ua
headerss = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36'
}



hen = 1
for i in lineslist:
    data = ws['A%d' % hen] = i #写入表格 的A列
    fullurl = url + i + url2 + i + url3
    html = requests.get(fullurl,headers=headerss) # 可以良心一下加个  timeout=5,
    htmlutf = html.content   #转码步骤一
    html_doc=str(htmlutf,'utf-8') #html_doc=html.decode("utf-8","ignore")   #转码步骤二 转换成utf8
    urls = re.findall('找到约 (.*?) 条结果',html_doc)   # <div id="result-stats">About (.*?) results<nobr> (0.47 seconds)&nbsp;</nobr></div>
    print(urls)
    for x in urls:
        data = ws['B%d' % hen] = x
        print('已添加'+i+x)
    hen = hen+1   #加一个值
    time.sleep(1)
wb.save('谷歌.xlsx')